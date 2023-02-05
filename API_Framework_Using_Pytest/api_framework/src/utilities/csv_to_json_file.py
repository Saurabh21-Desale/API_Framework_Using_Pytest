import csv
import json

csvFile = r'C:\Users\saurabhd\PycharmProjects\API_Automation_Practice\api_framework\test\test_data\csv_data.xlsx'
jsonFile = r'C:\Users\saurabhd\PycharmProjects\API_Automation_Practice\api_framework\test\test_data\json_data.json'

# class CsvToJsonFile(object):
#
#     def __init__(self):
#         pass
        # self.csvFilePath = r'C:\Users\saurabhd\PycharmProjects\API_Automation_Practice\api_framework\test\test_data\csv_data.xlsx'
        # self.jsonFilePath = r'C:\Users\saurabhd\PycharmProjects\API_Automation_Practice\api_framework\test\test_data\json_data.json'
from openpyxl import load_workbook
wb = load_workbook(csv)
from json import dumps

# get sheet by name
sheet = wb["student_data"]

# get number of rows
rows = sheet.max_row

# get number of columns
columns = sheet.max_column

# list to store all the rows of excel file as dictionary
lst = []
for i in range(1, rows):
    row = {}
    for j in range(1, columns):
        column_name = sheet.cell(row=1, column=j)
        row_data = sheet.cell(row=i+1, column=j)

        row.update(
            {
                column_name.value: row_data.value
            }
        )
    lst.append(row)


# convert into json
json_data = dumps(lst)
print(json_data)
